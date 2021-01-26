from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
import os

def unique(list1):
	list_set = set(list1)
	unique_list = (list(list_set))
	return unique_list

def build_invoice(cur_inv_lines):
	c = canvas.Canvas(f"output/{cur_inv_lines[0]['Number']}.pdf")
	aW = 540
	aH = 720
	counter = 0
	c.setLineWidth(.3)
	c.setFont('Helvetica-Bold', 12)


	# Page borders
	c.setLineWidth(2)
	c.line(15,10,15,830)
	c.line(15,830,580,830)
	c.line(580,830,580,10)
	c.line(580,10,15,10)


	# Logo
	c.drawImage("logo.png",170,740,width=280,height=80)

	# Company Info Table
	t1 = [['QAT Blended Learning Education ME', 'Tel: +971 2 650 5724'],
	      ['Apt. 203, Souq Tower', 'Email : qatc.info@gmail.com'],
	      ['Corniche Area, Abu Dhabi, UAE', 'TRN : 100552301200003']]
	table1 = Table(t1, colWidths=[aW/2,aW/2], rowHeights=[19,19,19])
	table1.setStyle(TableStyle([
 		('ALIGN', (1,1), (-1,-1), 'RIGHT'),
		("BACKGROUND", (0, 0), (-1, -1), '#D9E1F2'),
		("BOX", (0, 0), (-1, -1), 2, colors.black),
		('ALIGN', (0, 0), (-1, -1), "LEFT"),
		('LEFTPADDING', (0, 0), (-1, -1), 30),
		('VALIGN', (0, 0), (-1, -1), "MIDDLE")]))
	w, h = table1.wrap(aW, aH)
	table1.drawOn(c, 30, 675) #Table position

	# Tax Invoice Title
	c.setFont("Helvetica-Bold",14)
	c.drawString(aW/2-10,640,'TAX INVOICE')

	# Invoice Details
	c.setFont('Helvetica-Bold', 10)
	c.drawString(400,620,'Invoice No')
	c.drawString(480,620,':')
	c.drawString(490,620,str(cur_inv_lines[0]['Number']))
	c.drawString(400,600,'Invoice Date')
	c.drawString(480,600,':')
	c.drawString(490,600, cur_inv_lines[0]['Date'].strftime("%d / %b / %Y"))
	c.drawString(400,580,'Academic Year')
	c.drawString(480,580,':')
	c.drawString(490,580,str(cur_inv_lines[0]['Year']))

	# Student Information
	t2 = [
		['Student ID', ':', cur_inv_lines[0]['Student_No'],'1.','Total (Fees + VAT)',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Total_Fees'])],
	      ['Student Name', ':', cur_inv_lines[0]['Student_Name'],'2.','Previously Paid',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Previously_Paid'])],
	      ['Grade', ':', cur_inv_lines[0]['Grade'],'3.','Remaining Fees',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Remaining_Fees'])],
	      ['Education Type', ':', cur_inv_lines[0]['Education'],'4.','Invoices To Date',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Invoices_To_Date'])],
	      ['Parent Name',':',cur_inv_lines[0]['Parent_Name'],'5.','Balance Due',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Balance'])],
		  ['Parent Email',':',cur_inv_lines[0]['Parent_Email'],'','','','']
		]


	table2 = Table(t2, colWidths=[80,20,205,15,120,10,90], rowHeights=[17,17,17,17,17,17])
	table2.setStyle(TableStyle([
		("BACKGROUND", (0, 0), (-1, -1), '#D9E1F2'),
		("BOX", (0, 0), (-1, -1), 2, colors.black),
		('ALIGN', (0, 0), (-1, -1), "LEFT"),
		('LEFTPADDING', (0, 0), (-1, -1), 10),
		# ('SPAN', (3, 4), (4, 4)),
		('VALIGN', (0, 0), (-1, -1), "MIDDLE")]))
	w, h = table2.wrap(aW, aH)
	table2.drawOn(c, 30, 450) #Table position
	# Defaults Table
	defaults = [['Description', 'Amount', 'VAT 5%', 'Total'],
	        ['Registration Fees', '0','0', '0'],
	        ['Entrance Fees', '0', '0', '0'],
	        ['Monthly Installment - September', '0', '0', '0'],
	        ['Monthly Installment - October', '0', '0', '0'],
	        ['Monthly Installment - November', '0', '0', '0'],
	        ['Monthly Installment - December', '0', '0', '0'],
	        ['Monthly Installment - January', '0', '0', '0'],
	        ['Monthly Installment - February', '0', '0', '0'],
	        ['Monthly Installment - March', '0', '0', '0'],
	        ['Monthly Installment - April', '0', '0', '0'],
	        ['Monthly Installment - May', '0', '0', '0'],
	        ['Monthly Installment - June', '0', '0', '0'],
	        ['Total', '0', '0', '0']]
	items = ['Registration', 'Entrance', 'September', 'October', 'November', 'December', 'January', 'February', 'March', 'April', 'May', 'June', ]
	totals = [0,0,0]
	for i in range(len(cur_inv_lines)):	
		for x in range(len(defaults)-2):
			if cur_inv_lines[i]['Installment'] == items[x]:
				defaults[x+1][1] = '{:0,.2f}'.format(cur_inv_lines[i]['Invoice_Amount'])
				defaults[x+1][2] = '{:0,.2f}'.format(cur_inv_lines[i]['VAT'])
				defaults[x+1][3] = '{:0,.2f}'.format(cur_inv_lines[i]['Total'])
		totals[0] = totals[0] + cur_inv_lines[i]['Invoice_Amount']
		totals[1] = totals[1] + cur_inv_lines[i]['VAT']
		totals[2] = totals[2] + cur_inv_lines[i]['Total']
	defaults[13][1] = '{:0,.2f}'.format(totals[0])
	defaults[13][2] = '{:0,.2f}'.format(totals[1])
	defaults[13][3] = '{:0,.2f}'.format(totals[2])

	t = Table(defaults, colWidths=[240,100,100,100], rowHeights=[25,17,17,17,17,17,17,17,17,17,17,17,17,25])
	t.setStyle(TableStyle([
		("BOX", (0, 0), (-1, -1), 2, colors.black),
		('INNERGRID', (0, 0), (-1, -1), 1.5, colors.black),
		('ALIGN', (0, 0), (3, -1), "CENTER"),
		('ALIGN', (0, 1), (0, 12), "LEFT"),
		('LEFTPADDING', (0, 1), (0, 12), 30),
		('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
		('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
		('VALIGN', (0, 0), (-1, -1), "MIDDLE")]))
	w, h = t.wrap(aW, aH)
	t.drawOn(c, 30, 180) #Table position

	# Invoice Notes
	c.drawString(60,160,'This Invoice is due within 5 days from invoice date.')

	# Invoice Signatures
	c.setLineWidth(.3)
	c.setFont('Helvetica', 12)
	c.drawString(60,110,'Accountant')
	c.line(55,85,125,85)
	c.drawString(245,110,'Approved')
	c.line(240,85,310,85)
	c.drawString(445,110,'Received')
	c.line(440,85,510,85)

	# Invoice footer
	c.setFont('Helvetica', 12)
	c.drawString(80,30,'"This invoice is automatically generate and does not require stamp or signature"')
	c.save()

def build_credit_note(cur_inv_lines):
	c = canvas.Canvas(f"output/{cur_inv_lines[0]['Number']}.pdf")
	aW = 540
	aH = 720
	counter = 0
	c.setLineWidth(.3)
	c.setFont('Helvetica-Bold', 12)

	# Page borders
	c.setLineWidth(2)
	c.line(15,10,15,830)
	c.line(15,830,580,830)
	c.line(580,830,580,10)
	c.line(580,10,15,10)

	# Tax Credit Note Title
	c.setFont("Helvetica-Bold",14)
	c.drawString(aW/2-50,640,'TAX CREDIT NOTE')
	styles = getSampleStyleSheet()
	style = styles["BodyText"]



	# Logo
	c.drawImage("logo.png",170,740,width=280,height=80)

	# Company Info Table
	t1 = [['QAT Blended Learning Education ME', 'Tel: +971 2 650 5724'],
	      ['Apt. 203, Souq Tower', 'Email : qatc.info@gmail.com'],
	      ['Corniche Area, Abu Dhabi, UAE', 'TRN : 100552301200003']]
	table1 = Table(t1, colWidths=[aW/2,aW/2], rowHeights=[19,19,19])
	table1.setStyle(TableStyle([
		("BACKGROUND", (0, 0), (-1, -1), '#D9E1F2'),
		("BOX", (0, 0), (-1, -1), 2, colors.black),
		('ALIGN', (0, 0), (-1, -1), "LEFT"),
		('LEFTPADDING', (0, 0), (-1, -1), 30),
		('VALIGN', (0, 0), (-1, -1), "MIDDLE")]))
	w, h = table1.wrap(aW, aH)
	table1.drawOn(c, 30, 675) #Table position


	# Credit Note Details
	c.setFont('Helvetica-Bold', 10)
	c.drawString(400,620,'Credit Note No')
	c.drawString(480,620,':')
	c.drawString(490,620,str(cur_inv_lines[0]['Number']))
	c.drawString(400,600,'Credit Note Date')
	c.drawString(480,600,':')
	c.drawString(490,600, cur_inv_lines[0]['Date'].strftime("%d / %b / %Y"))
	c.drawString(400,580,'Aganist Invoice No')
	c.drawString(480,580,':')
	c.drawString(490,580, str(cur_inv_lines[0]['Against']))
	
	# Student Information
	t2 = [
	['Student ID', ':', cur_inv_lines[0]['Student_No'],'1.','Total (Fees + VAT)',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Total_Fees'])],
		['Student Name', ':', cur_inv_lines[0]['Student_Name'],'2.','Previously Paid',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Previously_Paid'])],
		['Grade', ':', cur_inv_lines[0]['Grade'],'3.','Remaining Fees',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Remaining_Fees'])],
		['Education Type', ':', cur_inv_lines[0]['Education'],'4.','Invoices To Date',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Invoices_To_Date'])],
		['Parent Name',':',cur_inv_lines[0]['Parent_Name'],'5.','Balance Due',':','AED {:0,.2f}'.format(cur_inv_lines[0]['Balance'])],
		['Parent Email',':',cur_inv_lines[0]['Parent_Email'],'','','','']
	]
	table2 = Table(t2, colWidths=[80,20,205,15,120,10,90], rowHeights=[17,17,17,17,17,17])
	table2.setStyle(TableStyle([
		("BACKGROUND", (0, 0), (-1, -1), '#D9E1F2'),
		("BOX", (0, 0), (-1, -1), 2, colors.black),
		('ALIGN', (0, 0), (-1, -1), "LEFT"),
		('LEFTPADDING', (0, 0), (-1, -1), 10),
		# ('SPAN', (3, 4), (4, 4)),
		('VALIGN', (0, 0), (-1, -1), "MIDDLE")]))
	w, h = table2.wrap(aW, aH)
	table2.drawOn(c, 30, 450) #Table position

	# Defaults Table
	defaults = [['Description', 'Amount', 'VAT 5%', 'Total'],
	        ['Registration Fees', '0','0', '0'],
	        ['Entrance Fees', '0', '0', '0'],
	        ['Monthly Installment - September', '0', '0', '0'],
	        ['Monthly Installment - October', '0', '0', '0'],
	        ['Monthly Installment - November', '0', '0', '0'],
	        ['Monthly Installment - December', '0', '0', '0'],
	        ['Monthly Installment - January', '0', '0', '0'],
	        ['Monthly Installment - February', '0', '0', '0'],
	        ['Monthly Installment - March', '0', '0', '0'],
	        ['Monthly Installment - April', '0', '0', '0'],
	        ['Monthly Installment - May', '0', '0', '0'],
	        ['Monthly Installment - June', '0', '0', '0'],
	        ['Total', '0', '0', '0']]
	items = ['Registration', 'Entrance', 'September', 'October', 'November', 'December', 'January', 'February', 'March', 'April', 'May', 'June', ]
	totals = [0,0,0]
	for i in range(len(cur_inv_lines)):	
		for x in range(len(defaults)-2):
			if cur_inv_lines[i]['Installment'] == items[x]:
				defaults[x+1][1] = '{:0,.2f}'.format(cur_inv_lines[i]['Invoice_Amount'])
				defaults[x+1][2] = '{:0,.2f}'.format(cur_inv_lines[i]['VAT'])
				defaults[x+1][3] = '{:0,.2f}'.format(cur_inv_lines[i]['Total'])
		totals[0] = totals[0] + cur_inv_lines[i]['Invoice_Amount']
		totals[1] = totals[1] + cur_inv_lines[i]['VAT']
		totals[2] = totals[2] + cur_inv_lines[i]['Total']
	defaults[13][1] = '{:0,.2f}'.format(totals[0])
	defaults[13][2] = '{:0,.2f}'.format(totals[1])
	defaults[13][3] = '{:0,.2f}'.format(totals[2])

	t = Table(defaults, colWidths=[240,100,100,100], rowHeights=[25,17,17,17,17,17,17,17,17,17,17,17,17,25])
	t.setStyle(TableStyle([
		("BOX", (0, 0), (-1, -1), 2, colors.black),
		('INNERGRID', (0, 0), (-1, -1), 1.5, colors.black),
		('ALIGN', (0, 0), (3, -1), "CENTER"),
		('ALIGN', (0, 1), (0, 12), "LEFT"),
		('LEFTPADDING', (0, 1), (0, 12), 30),
		('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
		('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
		('VALIGN', (0, 0), (-1, -1), "MIDDLE")]))
	w, h = t.wrap(aW, aH)
	t.drawOn(c, 30, 180) #Table position

	# Invoice Notes
	# c.drawString(60,160,'This Invoice is due within 5 days from invoice date.')

	# Invoice Signatures
	c.setLineWidth(.25)
	c.setFont('Helvetica', 12)
	c.drawString(60,110,'Accountant')
	c.line(55,85,125,85)
	c.drawString(245,110,'Approved')
	c.line(240,85,310,85)
	c.drawString(445,110,'Received')
	c.line(440,85,510,85)

	# Invoice footer
	c.setFont('Helvetica', 12)
	c.drawString(70,30,'"This Credit Note is automatically generate and does not require stamp or signature"')
	c.save()

# Set Global Variables
success = 0
failed = 0
current_Invoice = 0
data = []
numbers_col = []

# Loop the Excel sheet and collect data
wb = openpyxl.load_workbook('data.xlsx', data_only=True)
sheet = wb['Sheet1']
print("---------------------------")
print("Total Rows ",sheet.max_row)
print("Total Columns ",sheet.max_column)
print("---------------------------")
for row in range(sheet.max_row-1):
	elem = {}
	for col in range(sheet.max_column):
		elem[sheet.cell(1,col+1).value] = sheet.cell(row+2,col+1).value
	data.append(elem)
	numbers_col.append(elem["Number"])

invoices = unique(numbers_col)
for cur_invoice in invoices:
	cur_inv_lines = []
	for line in data:
		if line["Number"] == cur_invoice:
			cur_inv_lines.append(line)
	if cur_inv_lines[0]["Type"] == "Invoice":
		build_invoice(cur_inv_lines)
		print(f"Invoice #{cur_invoice} Created")
	elif cur_inv_lines[0]["Type"] == "Credit Note":
		build_credit_note(cur_inv_lines)
		print(f"Credit Note #{cur_invoice} Created")
	else:
		print("error")

os.system("pause")





